# your_app/views.py

import os
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime
import pandas as pd
import threading
import cv2
import xlrd
from xlutils.copy import copy as xl_copy
import numpy as np
import face_recognition
from PIL import Image

# Constants for file paths
# Define the directory where letters will be stored
LETTERS_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'letters')  
# Path to letters folder
if not os.path.exists(LETTERS_FOLDER):
    os.makedirs(LETTERS_FOLDER)


import os
from django.conf import settings

EXCEL_FILE_PATH = os.path.join(settings.MEDIA_ROOT, 'data', 'data.xls')
EXCEL_FILE_PATH1 = os.path.join(settings.MEDIA_ROOT, 'data', 'data1.xls')

import os

# Ensure MEDIA_ROOT is properly configured
EXCEL_FILE_PATH = os.path.join(settings.MEDIA_ROOT, 'data', 'data.xls')
print(EXCEL_FILE_PATH)  # Debugging: Check the file path

# Check if the file exists
if os.path.exists(EXCEL_FILE_PATH):
    print("File exists!")
else:
    print("File not found!")
# Ensure folders exist
os.makedirs(LETTERS_FOLDER, exist_ok=True)

# Home view to load class.html
def class_view(request):
    return render(request, 'face_recognition_app/class.html')

def face_view(request):
    return render(request, 'face_recognition_app/face.html')

def student_view(request):
    return render(request, 'face_recognition_app/student.html')

def admins_view(request):
    return render(request, 'face_recognition_app/admin.html')

def about_view(request):
    return render(request, 'face_recognition_app/about.html')

def adminlogin_view(request):
    return render(request, 'face_recognition_app/adminlogin.html')

def studentlogin_view(request):
    return render(request, 'face_recognition_app/studentlogin.html')
def index_view(request):
    return render(request, 'face_recognition_app/index.html')
# views.py


def request_status_view(request):
    return render(request, 'face_recognition_app/requeststatus.html')


# Attendance result view (returns results from the attendance system)
def attendance_result_view(request):
    return render(request, 'face_recognition_app/attendance_results.html')

# Attendance page (renders attendance form)
def attendance_view(request):
    return render(request, 'face_recognition_app/attendance.html')

# Students request view (form for students to submit requests)
def students_request_view(request):
    return render(request, 'face_recognition_app/studentsrequest.html')

# View for viewing attendance
def view_attendance(request):
    return render(request, 'face_recognition_app/viewattendance.html')

# View for viewing students
def view_students_view(request):
    return render(request, 'face_recognition_app/viewstudent.html')
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
import os
import json

# Path to the folder containing the text files

STATUS_FILE = 'status.json'

# Load or create the status file
if not os.path.exists(STATUS_FILE):
    with open(STATUS_FILE, 'w') as file:
        json.dump({}, file)

def read_status():
    with open(STATUS_FILE, 'r') as file:
        return json.load(file)


from django.shortcuts import render, redirect
from django.http import JsonResponse
import openpyxl
from openpyxl import load_workbook

import xlrd

import xlrd

def student_login_process(request):
    if request.method == "POST":
        # Get the username and password from the request
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Open the Excel file using xlrd
        try:
            wb = xlrd.open_workbook(EXCEL_FILE_PATH1)
            sheet = wb.sheet_by_index(0)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f"Error opening file: {str(e)}"})

        # Debugging: Print the username and password received
        print(f"Received username: {username}, password: {password}")

        # Iterate through the rows of the Excel sheet (starting from row 1 to skip header)
        for row_idx in range(1, sheet.nrows):
            row = sheet.row_values(row_idx)
            stored_username = row[0].strip()  # Assuming username is in the first column

            # Convert the password cell value to string and strip any surrounding whitespace
            stored_password = str(row[1]).strip()  # Assuming password is in the second column

            # Debugging: Print the stored username and password
            print(f"Checking: {stored_username}, {stored_password}")

            # Compare the username and password
            if stored_username == username and stored_password == password:
                # Return a successful response if the username and password match
                return JsonResponse({'success': True, 'message': 'Login successful!'})

        # If no match was found, return an error message
        return JsonResponse({'success': False, 'message': 'Invalid username or password'})


def list_files(request):
    try:
        files = [f for f in os.listdir(LETTERS_FOLDER) if f.endswith('.txt')]
        status_data = read_status()  # Load current statuses from status.json
        file_statuses = [{'name': f, 'status': status_data.get(f, 'pending')} for f in files]
        return JsonResponse(file_statuses, safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# View to get the content of a specific file
def get_file_content(request, filename):
    try:
        if filename.endswith('.txt'):
            file_path = os.path.join(LETTERS_FOLDER, filename)
            with open(file_path, 'r') as file:
                file_content = file.read()
            return HttpResponse(file_content, content_type='text/plain')
        else:
            return JsonResponse({"error": "Invalid file type"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

# View to update the status of a file
import json

def update_status(request, filename):
    try:
        if not filename.endswith('.txt'):
            return JsonResponse({"error": "Invalid file type"}, status=400)

        # Get the action from the request body (JSON)
        data = json.loads(request.body)
        action = data.get('action')

        if action not in ['accepted', 'rejected']:
            return JsonResponse({"error": "Invalid action"}, status=400)

        status_data = read_status()
        status_data[filename] = action
        write_status(status_data)

        return JsonResponse({"message": f"Status of {filename} updated to {action}"})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
def write_status(status_data):
    with open(STATUS_FILE, 'w') as file:
        json.dump(status_data, file, indent=4)
    print(f"Status updated: {status_data}")

from django.http import JsonResponse
from .models import File  # Assuming you have a File model to store file data

def get_file_status(request, file_name):
    try:
        file = File.objects.get(name=file_name)
        return JsonResponse({
            'name': file.name,
            'content': file.content,  # Assuming 'content' is the field storing file contents
            'status': file.status  # Return status (accepted/rejected)
        })
    except File.DoesNotExist:
        return JsonResponse({'error': 'File not found'}, status=404)

def get_file_list(request):
    files = File.objects.all()
    file_data = [{'name': file.name, 'status': file.status} for file in files]
    return JsonResponse(file_data, safe=False)
# If you're using threading, ensure the process properly starts and stops.
attendance_process = None
@csrf_exempt
def take_attendance_view(request):
    print("Received request to start attendance")  # Debugging line
    global attendance_process
    if attendance_process is None or not attendance_process.is_alive():
        # Start the attendance process in a separate thread
        attendance_process = threading.Thread(target=start_attendance)
        attendance_process.start()
        return JsonResponse({'message': 'Attendance process started.'})
    else:
        return JsonResponse({'message': 'Attendance process is already running.'}, status=400)



# View for leave letter submission
from django.http import JsonResponse  # Import JsonResponse

@csrf_exempt
def submit_letter_view(request):
    if request.method == 'POST':
        leave_letter_content = request.POST.get('leave_letter', '')
        filename = request.POST.get('filename', '')

        # Generate a timestamp-based filename if none is provided
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f'leave_letter_{timestamp}.txt'
        else:
            if not filename.endswith('.txt'):
                filename = f'{filename}.txt'

        # Save the letter content to a file
        file_path = os.path.join(LETTERS_FOLDER, filename)
        with open(file_path, 'w') as f:
            f.write(leave_letter_content)
        
        return HttpResponse(f'Letter has been successfully submitted and saved as {filename}!')
    return HttpResponse("Invalid request method.", status=400)


@csrf_exempt

def calculate_attendance(username):
    total_present = 0
    total_absent = 0

    # Load the Excel file using pandas
    df = pd.read_excel(EXCEL_FILE_PATH)

    # Print column names for debugging
    print("Columns in DataFrame:", df.columns)

    # Strip whitespace from column names
    df.columns = df.columns.str.strip()
    video_capture = cv2.VideoCapture(0)

    # Check if 'Username' exists in columns
    if 'Username' not in df.columns:
        raise KeyError("The column 'Username' was not found in the Excel file.")

    # Iterate through the DataFrame
    for index, row in df.iterrows():
        if row['Username'] == username:
            for key in row.index:
                if key not in ['Username', 'Name']:
                    # Count as "Present" or "Absent" based on cell value, including empty cells as "Absent"
                    if row[key] == 'Present':
                        total_present += 1
                    elif row[key] == 'Absent' or pd.isna(row[key]) or row[key] == '':
                        total_absent += 1

    return total_present, total_absent
import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
@csrf_exempt
def check_attendance(request):
    if request.method == "POST":
        # Get the username and password from the request
        username = request.POST.get('username').strip()
        password = request.POST.get('password').strip()

        # Open the Excel file
        try:
            wb = xlrd.open_workbook(EXCEL_FILE_PATH1)
            sheet = wb.sheet_by_index(0)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f"Error opening file: {str(e)}"})

        # Debugging: Print the username and password received
        print(f"Received username: {username}, password: {password}")

        # Flag to check if user is found
        is_valid_user = False

        # Iterate through the rows of the Excel sheet (starting from row 1 to skip header)
        for row_idx in range(1, sheet.nrows):
            row = sheet.row_values(row_idx)
            stored_username = str(row[0]).strip()
            stored_password = str(row[1]).strip()

            # Debugging: Print the stored username and password
            print(f"Checking: {stored_username}, {stored_password}")

            # Compare the username and password
            if stored_username == username and stored_password == password:
                is_valid_user = True
                print(f"Match found for user: {username}")
                break  # Stop the loop as we found the match

        # If valid user, calculate attendance
        if is_valid_user:
            total_present, total_absent = calculate_attendance(username)
            print(f"Attendance calculated - Present: {total_present}, Absent: {total_absent}")

            # Render the attendance results page with results
            context = {
                'total_present': total_present,
                'total_absent': total_absent,
                'user_not_found': total_present == 0 and total_absent == 0,
            }
            return render(request, 'face_recognition_app/attendance_results.html', context)

        # If no match was found, return an error message
        print("No matching user found. Invalid credentials.")
        return JsonResponse({'success': False, 'message': 'Invalid username or password'})

    # If request method is not POST
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

from django.views.decorators.csrf import csrf_exempt
# views.py
import threading
from django.http import JsonResponse
from datetime import datetime
import os
import xlrd
from xlutils.copy import copy as xl_copy
import cv2
import face_recognition
import numpy as np
from django.conf import settings

# Global flag to control attendance process
stop_attendance_flag = False

EXCEL_FILE_PATH = os.path.join(settings.MEDIA_ROOT, 'data', 'data.xls')


def start_attendance():
    global stop_attendance_flag
    stop_attendance_flag = False  # Reset flag each time attendance starts

    # Load attendance data and setup
    rb = xlrd.open_workbook(EXCEL_FILE_PATH, formatting_info=True)
    wb = xl_copy(rb)
    sheet = wb.get_sheet(0)
    original_sheet = rb.sheet_by_index(0)

    names_in_sheet = [original_sheet.cell_value(row, 0) for row in range(1, original_sheet.nrows)]
    col_index = original_sheet.ncols
    sheet.write(0, col_index, datetime.now().strftime("%Y-%m-%d %H:%M"))

    attendance_taken = set()

    # Load images and names dynamically from the photos directory
    image_dir = os.path.join(settings.MEDIA_ROOT, 'photos')
    known_face_encodings = []
    known_face_names = []

    for filename in os.listdir(image_dir):
        if filename.endswith('.jpg') or filename.endswith('.png'):
            image_path = os.path.join(image_dir, filename)
            image = face_recognition.load_image_file(image_path)
            encoding = face_recognition.face_encodings(image)
            if encoding:
                known_face_encodings.append(encoding[0])
                # Assume name is the file name without extension
                name = os.path.splitext(filename)[0]
                known_face_names.append(name)

    video_capture = cv2.VideoCapture(0)

    while not stop_attendance_flag:
        ret, frame = video_capture.read()
        if not ret:
            print("Failed to capture frame.")
            continue

        # Resize and convert frame for face recognition
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = small_frame[:, :, ::-1]

        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
        face_names = []

        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.4)
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            name = known_face_names[best_match_index] if matches[best_match_index] else "Unknown"
            face_names.append(name)

            if name in names_in_sheet and name not in attendance_taken:
                row_index = names_in_sheet.index(name) + 1
                sheet.write(row_index, col_index, "Present")
                attendance_taken.add(name)
                print(f"Attendance taken for {name}")
                wb.save(EXCEL_FILE_PATH)

        # Draw boxes and labels around recognized faces
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            # Draw a box around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            # Draw a label with a name below the face
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

        # Display the resulting image
        cv2.imshow('Video', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            print("Exiting attendance process.")
            break

    # Release video capture resources
    video_capture.release()
    cv2.destroyAllWindows()
    return JsonResponse({"status": "Attendance process stopped."})


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def stop_attendance(request):
    global stop_attendance_flag
    stop_attendance_flag = True  # Set the flag to stop the loop in start_attendance

    # Open the Excel sheet to mark absentees
    rb = xlrd.open_workbook(EXCEL_FILE_PATH, formatting_info=True)
    wb = xl_copy(rb)
    sheet = wb.get_sheet(0)
    original_sheet = rb.sheet_by_index(0)

    names_in_sheet = [original_sheet.cell_value(row, 0) for row in range(1, original_sheet.nrows)]
    col_index = original_sheet.ncols - 1  # Use the latest column

    # Mark absent for students who do not have "Present" marked
    for i, name in enumerate(names_in_sheet, start=1):
        if not original_sheet.cell_value(i, col_index):  # Check if cell is empty
            sheet.write(i, col_index, "Absent")
            print(f"Marked {name} as Absent")

    # Save the updated attendance sheet
    wb.save(EXCEL_FILE_PATH)

    return JsonResponse({"status": "Attendance process stopped, and absentees marked."})


import os
import xlrd
import xlwt
from xlutils.copy import copy
from django.conf import settings
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

EXCEL_FILE_PATH1 = os.path.join(settings.MEDIA_ROOT, 'data', 'data1.xls')
EXCEL_FILE_PATH = os.path.join(settings.MEDIA_ROOT, 'data', 'data.xls')

@csrf_exempt
def add_student(request):
    if request.method == 'POST':
        name = request.POST['name']
        username = request.POST['username']
        password = request.POST['password']
        photo = request.FILES['photo']

        # Save the photo as .jpg
        photo_path = os.path.join(settings.MEDIA_ROOT, 'photos', f"{username}.jpg")
        with open(photo_path, 'wb+') as destination:
            for chunk in photo.chunks():
                destination.write(chunk)

        # Add data to data1.xls (Username and Password)
        if os.path.exists(EXCEL_FILE_PATH1):
            rb = xlrd.open_workbook(EXCEL_FILE_PATH1, formatting_info=True)
            wb = copy(rb)
            sheet = wb.get_sheet(0)
            row = rb.sheet_by_index(0).nrows
            sheet.write(row, 0, username)
            sheet.write(row, 1, password)
            wb.save(EXCEL_FILE_PATH1)
        else:
            wb = xlwt.Workbook()
            sheet = wb.add_sheet('Sheet1')
            sheet.write(0, 0, 'Username')
            sheet.write(0, 1, 'Password')
            sheet.write(1, 0, username)
            sheet.write(1, 1, password)
            wb.save(EXCEL_FILE_PATH1)

        # Add data to data.xls (Name and Username)
        if os.path.exists(EXCEL_FILE_PATH):
            rb = xlrd.open_workbook(EXCEL_FILE_PATH, formatting_info=True)
            wb = copy(rb)
            sheet = wb.get_sheet(0)
            row = rb.sheet_by_index(0).nrows
            sheet.write(row, 0, name)
            sheet.write(row, 1, username)
            wb.save(EXCEL_FILE_PATH)
        else:
            wb = xlwt.Workbook()
            sheet = wb.add_sheet('Sheet1')
            sheet.write(0, 0, 'Name')
            sheet.write(0, 1, 'Username')
            sheet.write(1, 0, name)
            sheet.write(1, 1, username)
            wb.save(EXCEL_FILE_PATH)

        return JsonResponse({"status": "Student added successfully!"})

    return redirect('admin')
